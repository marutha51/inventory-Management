from concurrent.futures import ThreadPoolExecutor
import time
import openpyxl
import psycopg2

Awesome_Women_Affilation_Id = '7af20a4e-ca24-e010-34b6-bf119272c984'

GET_PRODUCT_QUERY="SELECT * FROM PRODUCTS WHERE id='{product_id}'"

GET_BLACKLISTED_AFFILIATION = '''SELECT product_id, string_agg(a.name, ', ') as blacklisted
    FROM product_affiliation_blacklists pab join affiliations a on pab.affiliation_id=a.id 
    WHERE product_id='{product_id}'
    GROUP BY product_id;'''

ADD_AFFILIATION_BLACKLIST_QUERY = '''
    INSERT INTO PRODUCT_AFFILIATION_BLACKLISTS(PRODUCT_ID,AFFILIATION_ID,ACTIVATION_DATE,CREATED_AT,UPDATED_AT)
    VALUES('{product_id}','{affiliation_id}',NOW(),NOW(),NOW())
'''

PROD_DB_PARAMS = {
    "host": "production-api-gateway-db-bc1020d911eab872.postgres.database.azure.com",
    "database": "api-gateway",
    "user": "maruthaprod@production-api-gateway-db-bc1020d911eab872.postgres.database.azure.com",
    "password":"b3TDz!VEp"
}

STAGE_DB_PARAMS = {
    "host": "staging-api-gateway-db-0c2946783cc4e1ce.postgres.database.azure.com",
    "database": "api-gateway",
    "user": "maruthastaging@staging-api-gateway-db-0c2946783cc4e1ce.postgres.database.azure.com",
    "password":"_Q3uv3XZ7"
}

FILE_LOCATION = 'Data.xlsx'
SHEET_NAME = 'CapRx'
# Open the Data sheet
work_book = openpyxl.load_workbook(FILE_LOCATION)
work_sheet = work_book[SHEET_NAME]

# New File for store the query execution status
result_book = openpyxl.Workbook()
result_sheet = result_book.active
result_sheet.append(['id', 'status', 'message','quey'])

def get_db_connection(params):
    conn = psycopg2.connect(**params)
    conn.autocommit = True
    return conn

def get_productid_column():
    headerRow = work_sheet[1]
    for title in headerRow:
        print(title.value)
        if title.value == 'product uuid':
            return title.column

def execute_product_get_query(product_id):
    if not product_id:
        return
    try:
        # query = GET_PRODUCT_QUERY.format(product_id=product_id)
        query = GET_BLACKLISTED_AFFILIATION.format(product_id=product_id)
        cur.execute(query)
        response = cur.fetchall()
        print(product_id, response)
        result_sheet.append([ product_id, 'success', str(response), query])
    except Exception as e:
        result_sheet.append([ product_id, 'Fail', str(e), query])

def add_aw_to_blacklist(product_id):
    if not product_id:
        return
    try:
        query = ADD_AFFILIATION_BLACKLIST_QUERY.format(product_id=product_id, affiliation_id=Awesome_Women_Affilation_Id)
        cur.execute(query.strip())
        result_sheet.append([ product_id, 'success', 'Inserted', query])
    except Exception as e:
        print('error',e)
        result_sheet.append([ product_id, 'Fail', str(e), query])
    
try:
    product_id_column = get_productid_column()
    print(product_id_column)
    db = get_db_connection(PROD_DB_PARAMS)
    product_uuid_list = []
    for col in work_sheet.iter_cols(min_row=11, max_row=3993, min_col=product_id_column, max_col=product_id_column, values_only=True):
        product_uuid_list = [ cell for cell in col]

    print('count:',len(product_uuid_list))
    with db.cursor() as cur:
        for id in product_uuid_list:
            add_aw_to_blacklist(id)
        
    result_book.save(f"Aff_blist_exec_report_{time.time()}.xlsx")
except (Exception, psycopg2.DatabaseError) as error:
        print('Error', error)
finally:
    db.close()



